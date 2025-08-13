"""
Data Loader for Online Retail II Dataset
Converts Excel to MySQL using SQLAlchemy for TASK 6: Sales Trend Analysis

This script:
1. Reads Online Retail II Excel file (all sheets)
2. Normalizes dates and handles missing columns
3. Loads data into MySQL database using SQLAlchemy
4. Creates the required table structure for analysis

Requirements:
- pandas, sqlalchemy, pymysql, openpyxl, python-dateutil, tqdm
"""

import argparse
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation

from dateutil import parser as dateparser
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from tqdm import tqdm
import pandas as pd


def normalize_column_name(name: str) -> str:
    """Normalize column names for consistent mapping"""
    return "".join(ch for ch in name.lower().strip() if ch.isalnum())


# Column mapping for different Excel header variations
COLUMN_CANDIDATES = {
    "invoice":      ["invoice", "invoiceno", "invoice no"],
    "stock_code":   ["stockcode", "stock code", "productid", "product id"],
    "description":  ["description", "productdescription", "product description"],
    "quantity":     ["quantity", "qty"],
    "invoice_date": ["invoicedate", "invoice date", "orderdate", "order date"],
    "unit_price":   ["unitprice", "unit price", "price"],
    "customer_id":  ["customerid", "customer id"],
    "country":      ["country", "region"],
}


def build_header_mapping(header_row):
    """Build mapping from expected columns to Excel column indices"""
    norm_to_idx = {normalize_column_name(str(h)): i for i, h in enumerate(header_row) if h is not None}
    result = {}
    
    for canonical_name, variants in COLUMN_CANDIDATES.items():
        idx = None
        for variant in variants:
            idx = norm_to_idx.get(normalize_column_name(variant))
            if idx is not None:
                break
        result[canonical_name] = idx
    
    return result


def safe_int(value):
    """Convert to int safely, return None for invalid values"""
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def safe_decimal(value):
    """Convert to Decimal safely, return None for invalid values"""
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        try:
            return Decimal(str(float(value)))
        except Exception:
            return None


def safe_string(value):
    """Convert to string safely, return None for empty values"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s != "" else None


def safe_datetime(value):
    """Convert to datetime safely, return None for invalid values"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    try:
        return dateparser.parse(str(value), dayfirst=True)
    except Exception:
        return None


def read_excel_sheets(file_path):
    """Generator that yields normalized records from all Excel sheets"""
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        
        try:
            header = next(rows)
        except StopIteration:
            continue
            
        header_map = build_header_mapping(header)

        # Check if this sheet has required columns
        required = ["invoice", "stock_code", "quantity", "invoice_date", "unit_price"]
        missing = [col for col in required if header_map.get(col) is None]
        if missing:
            print(f"Skipping sheet '{worksheet.title}' - missing columns: {missing}")
            continue

        print(f"Processing sheet: '{worksheet.title}'")
        
        for row in rows:
            def get_column_value(canonical_name):
                idx = header_map.get(canonical_name)
                return row[idx] if idx is not None and idx < len(row) else None

            yield {
                "invoice":     safe_string(get_column_value("invoice")),
                "stock_code":  safe_string(get_column_value("stock_code")),
                "description": safe_string(get_column_value("description")),
                "quantity":    safe_int(get_column_value("quantity")) or 0,
                "invoice_date": safe_datetime(get_column_value("invoice_date")),
                "unit_price":  safe_decimal(get_column_value("unit_price")) or Decimal("0"),
                "customer_id": safe_int(get_column_value("customer_id")),
                "country":     safe_string(get_column_value("country")),
            }


def setup_database(engine, database_name):
    """Create database and table structure"""
    with engine.begin() as conn:
        conn.execute(text(f"CREATE DATABASE IF NOT EXISTS `{database_name}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"))
        conn.execute(text(f"USE `{database_name}`"))
        
        # Drop and recreate table for clean import
        conn.execute(text("DROP TABLE IF EXISTS online_retail_raw"))
        conn.execute(text("""
            CREATE TABLE online_retail_raw (
              invoice       VARCHAR(32),
              stock_code    VARCHAR(64),
              description   TEXT,
              quantity      INT,
              invoice_date  DATETIME,
              unit_price    DECIMAL(10,2),
              customer_id   INT NULL,
              country       VARCHAR(64),
              INDEX idx_invoice_date (invoice_date),
              INDEX idx_invoice (invoice)
            )
        """))


def bulk_insert_data(engine, database_name, data_generator, batch_size=10000):
    """Insert data in batches for better performance"""
    insert_query = """
        INSERT INTO online_retail_raw
        (invoice, stock_code, description, quantity, invoice_date, unit_price, customer_id, country)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """
    
    total_inserted = 0
    batch = []
    
    raw_connection = engine.raw_connection()
    try:
        with raw_connection.cursor() as cursor:
            cursor.execute(f"USE `{database_name}`")
            
            for record in tqdm(data_generator, desc="Loading records", unit="rows", mininterval=1.0):
                batch.append((
                    record["invoice"],
                    record["stock_code"],
                    record["description"],
                    record["quantity"],
                    record["invoice_date"].strftime("%Y-%m-%d %H:%M:%S") if record["invoice_date"] else None,
                    str(record["unit_price"]),
                    record["customer_id"],
                    record["country"],
                ))
                
                if len(batch) >= batch_size:
                    cursor.executemany(insert_query, batch)
                    raw_connection.commit()
                    total_inserted += len(batch)
                    batch.clear()
            
            # Insert remaining records
            if batch:
                cursor.executemany(insert_query, batch)
                raw_connection.commit()
                total_inserted += len(batch)
                
    finally:
        raw_connection.close()
    
    return total_inserted


def main():
    parser = argparse.ArgumentParser(description="Load Online Retail II data into MySQL")
    parser.add_argument("--file", required=True, help="Path to online_retail_II.xlsx")
    parser.add_argument("--user", default="root", help="MySQL username")
    parser.add_argument("--password", default="", help="MySQL password")
    parser.add_argument("--host", default="127.0.0.1", help="MySQL host")
    parser.add_argument("--port", type=int, default=3306, help="MySQL port")
    parser.add_argument("--database", default="online_sales", help="Database name")
    parser.add_argument("--batch-size", type=int, default=10000, help="Batch size for inserts")
    
    args = parser.parse_args()

    if not os.path.exists(args.file):
        raise SystemExit(f"File not found: {args.file}")

    # Create SQLAlchemy engine
    connection_url = f"mysql+pymysql://{args.user}:{args.password}@{args.host}:{args.port}/?charset=utf8mb4"
    engine = create_engine(connection_url, pool_pre_ping=True, future=True)

    print(f"Setting up database: {args.database}")
    setup_database(engine, args.database)
    
    print(f"Reading Excel file: {args.file}")
    data_stream = read_excel_sheets(args.file)
    
    print("Inserting data into MySQL...")
    total_records = bulk_insert_data(engine, args.database, data_stream, batch_size=args.batch_size)

    print(f"\n✅ Import completed successfully!")
    print(f"📊 Total records inserted: {total_records:,}")
    print(f"🗄️  Database: {args.database}")
    print(f"📋 Table: online_retail_raw")
    print("\nYou can now run the SQL analysis queries from TASK6_Sales_Trend_Analysis.sql")


if __name__ == "__main__":
    main()
